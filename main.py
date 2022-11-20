import re
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from multiprocessing import Pool
import aux_functions
from bs4 import BeautifulSoup
import time
import json
import random
import os
from dotenv import load_dotenv, find_dotenv
from collect_data import CollectData


def get_company_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        file_data = file.read()
        src = BeautifulSoup(file_data, 'lxml')
        data_collect = CollectData(src)
        data = data_collect.get_card_info()
        data['linkedin_url'] = get_linkedin_url(file_path)
        return data


def get_linkedin_url(file_path: str):
    file_numbers = re.findall('[0-9]+', file_path)
    with open(f'companies/test_{file_numbers[1]}.json', 'r', encoding='utf-8') as file:
        python_dict = json.loads(file.read())
        return python_dict[int(file_numbers[0]) - 1]


def save_company_src(driver, python_dict, json_file_numb):
    for num, each in enumerate(python_dict):
        files = [os.path.join('companies_page_src', filename) for filename in os.listdir('companies_page_src')]
        if f'companies_page_src\\company_{num + 1}_from_json_{json_file_numb}.html' in files:
            continue
        url = f'{each}about/'
        driver.get(url)
        time.sleep(random.randint(2, 4))
        with open(f'companies_page_src/company_{num + 1}_from_json_{json_file_numb}.html', 'w',
                  encoding='utf-8') as file:
            file.write(driver.page_source)
            print(f'-> company_{num + 1}_from_json_{json_file_numb} saved')


def find_companies_page_src(driver, json_files):
    for json_file_numb in range(json_files - 19, json_files + 1):
        with open(f'companies/test_{json_file_numb}.json', 'r', encoding='utf-8') as file:
            json_file = file.read()
            python_dict = json.loads(json_file)
            save_company_src(driver, python_dict, json_file_numb)


def find_companies_url(driver, pages):
    aux_func = aux_functions.AuxFunc(driver)
    for i in range(pages - 19, pages + 1):
        url = f'https://www.linkedin.com/search/results/companies/?heroEntityKey=urn%3Ali%3Aautocomplete%3A1554627024&' \
              f'keywords=crypto%20mining&origin=SWITCH_SEARCH_VERTICAL&page={i}&position=0&' \
              f'searchId=689af0f0-8580-418f-a909-48cbfc5ad764&sid=%3A3P'
        driver.get(url)
        time.sleep(random.randint(2, 4))
        company_url_xpath = '//*[@class[contains(.,"entity-result__title-text")]]//*[@class="app-aware-link "]'
        company_urls_element = aux_func.try_get_elements(company_url_xpath)
        company_urls = [url.get_attribute('href') for url in company_urls_element]
        with open(f'companies/test_{i}.json', 'w', encoding='utf-8') as file:
            file.write(json.dumps(company_urls))
        print(f'Page {i} parsed')


def init_driver():
    s = Service('C:/PyProject/chromedriver.exe')
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    browser = webdriver.Chrome(service=s)  # , options=chrome_options)
    return browser


def login(driver):
    url = 'https://www.linkedin.com/home'
    driver.get(url)
    time.sleep(5)
    load_dotenv(find_dotenv())
    linkedin_password = os.getenv('PASSWORD')
    linkedin_login = os.getenv('LOGIN')
    driver.find_element(By.XPATH, '//*[@name="session_key"]').send_keys(linkedin_login)
    driver.find_element(By.XPATH, '//*[@name="session_password"]').send_keys(linkedin_password)
    driver.find_element(By.XPATH, '//*[@class="sign-in-form__submit-button"]').click()
    time.sleep(5)


def get_companies_urls(pages):
    driver = init_driver()
    login(driver)
    find_companies_url(driver, pages)
    driver.close()
    driver.quit()


def get_companies_page_src(json_files_amount):
    driver = init_driver()
    login(driver)
    find_companies_page_src(driver, json_files_amount)
    driver.close()
    driver.quit()


def multiprocessing_get_url():
    # collect all companies URL
    pages_numb = [20, 40, 60, 80, 100]
    p = Pool(processes=len(pages_numb))
    p.map(get_companies_urls, pages_numb)


def multiprocessing_get_companies_src():
    # collect all companies URL
    json_files_amount = [20, 40, 60, 80, 100]
    # json_files_amount = [10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
    # json_files_amount = [1]
    p = Pool(processes=len(json_files_amount))
    p.map(get_companies_page_src, json_files_amount)


def collect_data():
    files = [os.path.join('companies_page_src', filename) for filename in os.listdir('companies_page_src')]
    all_data = []
    for num, file in enumerate(files):
        all_data.append(get_company_data(file))
        print(f'{num + 1} / {len(files)} company parsed')
    with open('crypto_companies_data.json', 'w', encoding='utf-8') as file:
        file.write(json.dumps(all_data))


def write_to_excel():
    with open('crypto_companies_data.json', 'r', encoding='utf-8') as file:
        python_dict = json.loads(file.read())

        workbook = load_workbook(filename='miners_data.xlsx')
        sheet = 'data'
        column = 1
        excel_row = 1
        for key, value in python_dict[0].items():
            workbook[sheet].cell(row=excel_row, column=column).value = key
            column += 1
        excel_row = 2
        for data_dict in python_dict:
            for num, each in enumerate(data_dict):
                workbook[sheet].cell(row=excel_row, column=num + 1).value = data_dict[each]
            excel_row += 1
        workbook.save('miners_data.xlsx')
        workbook.close()


if __name__ == '__main__':
    # multiprocessing_get_url()
    # multiprocessing_get_companies_src()
    # collect_data()
    write_to_excel()
    pass
